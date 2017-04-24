"""
Generate TSV Fiels

"""

from __future__ import division
from __future__ import print_function


# Import OS Functions
import argparse

# from openpyxl import Workbook
from openpyxl import load_workbook

import datetime
import re

# import decimal
# import textwrap
# import time
# import math

# import numpy as np
# import pandas as pd

import acc_lib as alib


# --------------------------------------------------------------------
#
#                          constants
#
# --------------------------------------------------------------------

CLUB_LIST = alib.CLUB_LIST
CLUB_TAGS = alib.CLUB_TAGS
OTHER_TAGS = alib.OTHER_TAGS

# ------------------------------------------------------------------------------------------
#
#                                       PRE PROCESS source
#
# ------------------------------------------------------------------------------------------


def clean_cell(p_cell, p_cell_location=None):
    """
    Remove non-ascii characters from cell

    Parameters
    ----------
    a cell value

    Returns
    -------
    None if no change
    modified value if there is a change

    """
    def clean_text(p_field):
        """
        Lambda function to replace characters that bryte cannot tolerate
        """
        dodgy_char = '’'
        dodgy_char2 = '‘'
        dodgy_char3 = '–'
        dodgy_char4 = '”'
        dodgy_char5 = '“'
        dodgy_char6 = '…'
        dodgy_char7 = '\r'
        dodgy_char8 = '\n'
        dodgy_char10 = '\t'
        dodgy_char12 = '–'
        dodgy_char15 = '‐'      # This is a different hyphon

        # Every field is terminated with it.
        dodgy_char20 = '\xa0'   # this is some kind of weird space, but is non ascii.

        dodgy_char21 = '·'      # another bizarre hyphon
        dodgy_char22 = '│'
        dodgy_char23 = '√'

        if(p_field is None or
           isinstance(p_field, datetime.date) or
           not isinstance(p_field, str)):
            new_field = p_field
        else:

            new_field = re.sub(dodgy_char, "'", p_field)
            new_field = re.sub(dodgy_char2, "'", new_field)
            new_field = re.sub(dodgy_char3, "-", new_field)
            new_field = re.sub(dodgy_char4, '"', new_field)
            new_field = re.sub(dodgy_char5, '"', new_field)
            new_field = re.sub(dodgy_char6, '.', new_field)
            new_field = re.sub(dodgy_char7, ' ', new_field)
            new_field = re.sub(dodgy_char8, ' ', new_field)
            new_field = re.sub(dodgy_char10, ' ', new_field)
            new_field = re.sub(dodgy_char12, '-', new_field)
            new_field = re.sub(dodgy_char15, '-', new_field)

            new_field = re.sub(dodgy_char20, '', new_field)
            new_field = re.sub(dodgy_char21, '-', new_field)
            new_field = re.sub(dodgy_char22, '|', new_field)
            new_field = re.sub(dodgy_char23, ' ', new_field)

            new_field = re.sub('[^ -~]', ' ', new_field)

            # remove trailing spaces
            new_field = re.sub(' *$', '', new_field)

        return new_field

    result = clean_text(p_cell)

    if result == p_cell:
        return None
    else:
        # alib.log_info('        before: {}'.format(p_cell))
        if p_cell_location is None:
            alib.log_info('        loc: , modified value : {}'.format(p_cell_location, result))
        else:
            alib.log_info('        loc: {}, modified value : {}'.format(p_cell_location, result))
        return result

    return


# --------------------------------------------------------------------
#
#                          create dir
#
# --------------------------------------------------------------------

def vh_files(p_list):
    """ find hidden sheets """

    for f in p_list:
        file_mod = False
        short_name = '/'.join(f.split('/')[-5:])
        alib.p_i('')
        alib.p_i('Process file : {}'.format(short_name))

        wb = load_workbook(filename=f.replace('/', '\\'))
        sheet_names = wb.get_sheet_names()

        for sheet in sheet_names:
            ws = wb[sheet]

            if sheet in ('Version History',
                         'Configuration',
                         'Database Schema',
                         'Configuration Screens'):
                alib.p_i('    Skip worksheet {}'.format(sheet))
                continue

            if ws.sheet_state != 'visible':
                alib.p_e('    Skip worksheet {}, not visible'.format(sheet))
                continue
            alib.p_i('    Process worksheet {}'.format(sheet))
            for row in ws.iter_rows():
                for cell in row:
                    res = clean_cell(cell.value, cell.coordinate)
                    if res is None:
                        continue
                    else:
                        file_mod = True
                        cell.value = res
        if file_mod:
            alib.p_i('    file modified')
            new_filename = f.replace('.', '_mod.')
            wb.save(filename=new_filename)
        else:
            alib.p_i('    file NOT modified')

# --------------------------------------------------------------------
#
#                          create dir
#
# --------------------------------------------------------------------


def validate_hidden(p_files):
    """ find hidden sheets """

    vh_files(p_files['c'])
    vh_files(p_files['cc'])
    # vh_files(p_files['m'])
    # vh_files(p_files['mc'])

# --- Program Init
# --------------------------------------------------------------------
#
#                          initialise
#
# --------------------------------------------------------------------
#


def initialise():
    """
    Necessary initialisations for command line arguments
    """
    # Logfile for logging
    log_filename = alib.log_filename_init()
    if log_filename is None:
        print("\nError: Failed to initialise Log File Name. aborting\n")
        return alib.FAIL_GENERIC

    parser = argparse.ArgumentParser(description="""
     Example command lines:

    -d DEBUG --ss c:/tmp/x.xlsx
    --ss "C:/work/sample doc/Out of Service Codes.xlsx"
          """, formatter_class=argparse.RawTextHelpFormatter)

    # --- DB parameters ---
#    def_dir = 'C:/work/stuff/'
#    m_def_dir = 'master_OneDrive_2017-04-07/Master Validated Templates by Club (Controlled)'
#    mc_def_dir = 'master_common_OneDrive_2017-04-07/Master Validated Common Data Templates (Controlled)'

    parser.add_argument('--club_dir',
                        help='club files directory',
                        default=None,
                        required=False)

    parser.add_argument('--club_common_dir',
                        help='club common files directory',
                        default=None,
                        required=False)

    parser.add_argument('--m_dir',
                        help='master directory',
                        default=None,
                        required=False)

    parser.add_argument('--mc_dir',
                        help='master common directory',
                        default=None,
                        required=False)

    parser.add_argument('--all_dir',
                        help='directory containing all of the above',
                        default=None,
                        required=False)

    parser.add_argument('--quick_debug',
                        help='only load cc and mc files, used for debugging only',
                        default=None,
                        action='store_true',
                        required=False)

    # Add debug arguments
    parser.add_argument('-d', '--debug',
                        help='Log messages verbosity: NONE (least), DEBUG (most)',
                        choices=('NONE', 'CRITICAL', 'ERROR', 'WARNING', 'INFO', 'DEBUG'),
                        default="INFO",
                        required=False)

    # Sort though the arguments, ensure mandatory are populated
    args = alib.args_validate(parser, log_filename)

    return (args, log_filename)

# --------------------------------------------------------------------
#
#                          main
#
# --------------------------------------------------------------------


def main():
    """
    This program tests that the classification document, the Sql Server DB and Redshift DB
    all agree with each other
    """

    args, dummy_l_log_filename_s = initialise()

    # -- Initialise
    if not alib.init_app(args):
        return alib.FAIL_GENERIC

    work_dir = alib.load_dir(args)
    print('Loading files from {}'.format(work_dir['l_c_dir']))
    work_files = alib.load_files(work_dir, args['quick_debug'])

    #    work_dict = alib.load_matching_masterfile(work_files)
    #    alib.load_tags(work_dict)
    #    alib.print_filenames(work_dict)

    validate_hidden(work_files)

    alib.p_i('Done...')

# --------------------------------------------------------------------
#
#                          the end
#
# --------------------------------------------------------------------

if __name__ == "__main__":
    if main() == alib.SUCCESS:
        exit
    else:
        exit(-1)

# --- eof ---
